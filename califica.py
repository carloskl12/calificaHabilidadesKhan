#!/usr/bin/env python
# -*- coding: utf-8 -*-

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
# Existencia de archivos
import os.path
import re
import time

wbDatos = load_workbook('datos.xlsx', read_only=True)


#Extrae los diferentes grupos existentes
patron=re.compile(r'\D+(\d+)$')
ws1=wbDatos['grupos']
grupos=[]
for cell in ws1[1]:
  coincidencia= patron.match(cell.value)
  numero_grupo= -1
  if coincidencia != None:
    numero_grupo= int( coincidencia.group(1) )
  else:
    print 'Error: se debe especificar bien el nombre del grupo \"%s\"'%cell.value
  grupos.append( numero_grupo )
#print 'Grupos: '+ str(grupos)

#Extrae las notas correspondientes para cada nivel de dominio
ws1=wbDatos[u'calificación']
nivel_dominio={}
for row in ws1['A2:B7']:
  key = row[0].value
  value = row[1].value
  if key != None and value !=None :
    nivel_dominio[key]=value
  else:
    print u'Error: están mal especificados los datos para calificación'
    break


x=u'Necesita práctia'
if nivel_dominio.has_key(x):
  print u'Sí existe la clave:%s'%x
else:
  print u'No existe la clave:%s'%x

ws1= wbDatos['datos']
numReportes= ws1.max_row - 1
for revision in range(numReportes):
  #%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%
  # Se repite para cada grupo diferente a evaluar
  ws1= wbDatos['datos']
  #Número de corte y grupo evaluado
  
  corte_evaluado = int(ws1.cell(row=revision+2,column=2).value)
  grupo_evaluado = int(ws1.cell(row=revision+2,column=3).value)
  print 'Grupo evaluado: %i  Corte:%i'%(grupo_evaluado,corte_evaluado)
  ub_grupo_evaluado=0;
  index=1
  for grupo in grupos:
    if grupo == grupo_evaluado:
      ub_grupo_evaluado= index
      break
    index+=1

  if ub_grupo_evaluado == 0:
    print 'No se halló el grupo a evaluar'
    #Acá se debe finalizar el script
    continue
  else:
    print 'Columna del grupo: '+ str(ub_grupo_evaluado)
  #Codigos de estudiantes

  ws1 = wbDatos['grupos']
  estudiantes = []

  rango= get_column_letter(ub_grupo_evaluado)
  rango+='2:'+rango+str(ws1.max_row)

  for cell in ws1[rango]:
    valor = cell[0].value
    if valor != None:
      estudiantes.append(str(valor))
    else:
      break

  print 'Total estudiantes: %i'% len(estudiantes)

  #Obtiene las habilidades
  ws1=wbDatos['habilidades']
  habilidades=[]
  rango= get_column_letter(corte_evaluado)
  rango+='2:'+rango+str(ws1.max_row)
  for cell in ws1[rango]:
    valor = cell[0].value
    if valor != None:
      habilidades.append(valor)
    else:
      break

  print 'Total de habilidades: '+ str(len(habilidades))

  ###################################################
  ############## Crea los diccionarios ##############
  # Para los estudiantes cada código lo asocia con una fila del formulario
  # Para las habilidades se asocian con una columna del formulario
  dic_estudiantes= { k:v for k,v  in zip (estudiantes, range(2,len(estudiantes)+2)) }
  dic_habilidades= {k:v for k,v in zip (habilidades, [get_column_letter(i+2) for i in range( len(habilidades))])}

  print u'Columna de habilidad: %s'%dic_habilidades[habilidades[25]]
  # Archivo fuente del reporte 
  ws1= wbDatos['datos']
  source = ws1.cell(row=revision+2 , column =1 ).value
  print "Archivo fuente: \"%s\""% source

  if os.path.exists(source):
    #Realiza la lectura del fichero
    wbReport= load_workbook(source)
    #wbReport= load_workbook(source,read_only=True)
    wsR= wbReport['Exercises']
    wbCalfica= Workbook()
    ws2= wbCalfica.active
    ws2['A1']=u'Códigos'
    start_time = time.time()
    for i in range(len(estudiantes)):
      ws2.cell(row=i+2,column=1,value=estudiantes[i])
    for i in range(len(habilidades)):
      ws2.cell(row=1,column=i+2,value=habilidades[i])
    
    for i in range(wsR.max_row):
      codigo_estudiante= wsR.cell(row=i+2,column=1).value
      habilidad_practicada= wsR.cell(row=i+2,column=2).value
      if not (type(codigo_estudiante) is str ):
        codigo_estudiante = str(codigo_estudiante)
      
      if dic_habilidades.has_key(habilidad_practicada) and dic_estudiantes.has_key(codigo_estudiante):
        dir_nota = dic_habilidades[habilidad_practicada]+ str(dic_estudiantes[codigo_estudiante])
        estado = wsR.cell(row=i+2,column=4).value
        ws2[dir_nota]= nivel_dominio[estado]

    
    print "--- %s seconds---" % (time.time() - start_time)
    
    wbCalfica.save(u'Califcación Grupo %i.xlsx'%grupo_evaluado)
  else:
    print "El fichero \"%s\" no existe"% source

print u'Se finalizó de revisar los reportes'

